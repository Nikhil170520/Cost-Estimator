import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as px
import numpy as np
from fpdf import FPDF
import openpyxl
from sklearn.linear_model import LinearRegression
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
import tempfile

# Set Streamlit page config
st.set_page_config(page_title="Cost Estimator", page_icon="💰", layout="wide")

# Apply custom CSS for better styling
st.markdown(
    """
    <style>
        .stApp {
            background-color: #1e1e2f;
            color: white;
        }
        .css-18e3th9 {
            padding-top: 2rem;
        }
        .stButton>button {
            background-color: #ff9800;
            color: white;
            border-radius: 10px;
        }
        h1, h2, h3, h4, h5, h6 {
            color: #ffcc00;
        }
        input[type=number], input[type=text] {
            padding: 10px;
            border-radius: 5px;
            border: 1px solid #ffcc00;
            background-color: #2a2a3a;
            color: white;
        }
        .stTextInput>div>div>input, .stNumberInput>div>div>input {
            background-color: #2a2a3a !important;
            color: white !important;
            border-radius: 5px;
            padding: 10px;
            border: 1px solid #ffcc00;
        }
    </style>
    """, unsafe_allow_html=True
)

# Function to calculate project cost
def calculate_cost(labor_cost, material_cost, equipment_cost, misc_cost):
    total_cost = labor_cost + material_cost + equipment_cost + misc_cost
    return total_cost

# Function to predict future costs
def predict_future_cost(data, years=5):
    if len(data) < 2:
        return "Not enough data for prediction"
    
    X = np.array(data["Year"]).reshape(-1, 1)
    y = np.array(data["Cost"]).reshape(-1, 1)
    
    model = LinearRegression()
    model.fit(X, y)
    
    future_years = np.array([data["Year"].max() + i for i in range(1, years+1)]).reshape(-1, 1)
    future_costs = model.predict(future_years)
    
    return future_years.flatten(), future_costs.flatten()

# Streamlit UI
st.title("💰 Cost Estimator")
st.sidebar.header("🔧 Input Project Details")

# User Inputs
project_name = st.text_input("Project Name", "", placeholder="Enter project name")
project_duration = st.number_input("Project Duration (in months)", min_value=1, step=1)
labor_cost = st.sidebar.number_input("Labor Cost", min_value=0, value=5000)
material_cost = st.sidebar.number_input("Material Cost", min_value=0, value=3000)
equipment_cost = st.sidebar.number_input("Equipment Cost", min_value=0, value=2000)
misc_cost = st.sidebar.number_input("Miscellaneous Cost", min_value=0, value=1000)

# Calculate total cost
total_cost = calculate_cost(labor_cost, material_cost, equipment_cost, misc_cost)
st.subheader(f"💸 Total Estimated Cost: INR {total_cost}")

# Cost Breakdown Chart
data = {"Category": ["Labor", "Material", "Equipment", "Miscellaneous"],
        "Cost": [labor_cost, material_cost, equipment_cost, misc_cost]}
df = pd.DataFrame(data)
fig = px.pie(df, names='Category', values='Cost', title='Cost Breakdown', color_discrete_sequence=px.colors.sequential.Plasma)
st.plotly_chart(fig, use_container_width=True)

# Historical Cost Data for Prediction (Example Data)
historical_data = pd.DataFrame({"Year": [2020, 2021, 2022, 2023, 2024],
                                "Cost": [10000, 12000, 14000, 16000, 18000]})

# Predict Future Costs
future_years, future_costs = predict_future_cost(historical_data)
st.subheader("📈 Cost Forecast for the Next 5 Years")
if isinstance(future_costs, str):
    st.write(future_costs)
else:
    pred_df = pd.DataFrame({"Year": future_years, "Predicted Cost": future_costs})
    fig_forecast = px.line(pred_df, x="Year", y="Predicted Cost", title="Future Cost Projection",
                           markers=True, line_shape='spline', template="plotly_dark")
    st.plotly_chart(fig_forecast, use_container_width=True)

# Generate and Download Report
if st.button("📄 Generate PDF Report"):
    if total_cost > 0:
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        pdf.cell(200, 10, "Project Cost Estimation Report", ln=True, align="C")
        pdf.ln(10)
        
        pdf.cell(0, 10, f"Project Name: {project_name}", ln=True)
        pdf.cell(0, 10, f"Project Duration: {project_duration} months", ln=True)
        pdf.cell(0, 10, f"Labor Cost: INR {labor_cost}", ln=True)
        pdf.cell(0, 10, f"Material Cost: INR {material_cost}", ln=True)
        pdf.cell(0, 10, f"Equipment Cost: INR {equipment_cost}", ln=True)
        pdf.cell(0, 10, f"Miscellaneous Cost: INR {misc_cost}", ln=True)
        pdf.cell(0, 10, f"Total Project Cost: INR {total_cost}", ln=True)

        temp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        pdf.output(temp_pdf.name)
        
        with open(temp_pdf.name, "rb") as pdf_file:
            st.download_button("📥 Download PDF", pdf_file, "Project_Cost_Report.pdf", mime="application/pdf")
    else:
        st.error("Total cost cannot be zero. Please enter valid values.")
